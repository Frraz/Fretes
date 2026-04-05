"""
python manage.py importar_planilha
Baixa a planilha do Google Drive e importa para o banco.
"""
import io
import re
import urllib.request
import urllib.error
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction
from core.models import Motorista, Romaneio, Abastecimento, Adiantamento, ImportLog


def _dec(val, default=Decimal("0")):
    if val is None: return default
    try:
        d = Decimal(str(val))
        return d if d.is_finite() else default
    except (InvalidOperation, ValueError, TypeError):
        return default

def _str(val):
    if val is None: return ""
    return str(val).strip()

def _date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    try: return datetime.strptime(str(val)[:10], "%Y-%m-%d").date()
    except: return None


def download_gdrive(file_id):
    """Baixa arquivo do Google Drive. Lida com confirmação de virus scan."""
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    headers = {"User-Agent": "Mozilla/5.0"}
    req = urllib.request.Request(url, headers=headers)
    resp = urllib.request.urlopen(req, timeout=60)
    data = resp.read()

    if b"<!DOCTYPE html>" in data[:500]:
        url2 = f"https://drive.google.com/uc?export=download&confirm=t&id={file_id}"
        req2 = urllib.request.Request(url2, headers=headers)
        resp2 = urllib.request.urlopen(req2, timeout=60)
        data = resp2.read()

    return data


class Command(BaseCommand):
    help = "Baixa planilha do Google Drive e importa para o banco"

    def add_arguments(self, parser):
        parser.add_argument("--file", type=str, help="Arquivo local (ignora Google Drive)")

    def handle(self, *args, **options):
        log = ImportLog()
        local_file = options.get("file")

        if local_file:
            self.stdout.write(f"\n  📂 Arquivo local: {local_file}")
            try:
                wb = openpyxl.load_workbook(local_file, data_only=True, read_only=True)
            except Exception as e:
                return self._fail(log, str(e))
        else:
            gdrive_id = getattr(settings, "GDRIVE_FILE_ID", "")
            if not gdrive_id:
                return self._fail(log, "GDRIVE_FILE_ID não configurado")

            self.stdout.write(f"\n  ☁️  Baixando do Google Drive...")
            try:
                data = download_gdrive(gdrive_id)
                self.stdout.write(f"  ✓ Download OK ({len(data)//1024} KB)")
                wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
            except Exception as e:
                return self._fail(log, f"Erro download: {e}")

        self.stdout.write(f"  📋 Abas: {wb.sheetnames}")
        sheet_map = {s.lower(): s for s in wb.sheetnames}
        cache = {}

        def _mot(nome, placa):
            key = (nome.upper().strip(), placa.upper().strip())
            if key not in cache:
                obj, _ = Motorista.objects.get_or_create(nome=key[0], placa=key[1])
                cache[key] = obj
            return cache[key]

        def _sheet(keyword):
            for k, v in sheet_map.items():
                if keyword in k: return wb[v]
            return None

        with transaction.atomic():
            Romaneio.objects.all().delete()
            Abastecimento.objects.all().delete()
            Adiantamento.objects.all().delete()

            rc = 0
            for gc, kw in [("SF", "romaneios sf"), ("SJ", "romaneios sj")]:
                ws = _sheet(kw)
                if not ws: continue
                bulk = []
                for r in range(3, ws.max_row + 1):
                    placa = _str(ws.cell(r, 8).value)
                    if not placa: continue
                    bulk.append(Romaneio(
                        grupo=gc, data=_date(ws.cell(r, 1).value),
                        nota_fiscal=_str(ws.cell(r, 2).value),
                        ticket=_str(ws.cell(r, 3).value),
                        origem=_str(ws.cell(r, 4).value),
                        talhao=_str(ws.cell(r, 5).value),
                        destino=_str(ws.cell(r, 6).value),
                        motorista=_mot(_str(ws.cell(r, 7).value), placa),
                        peso_liquido=_dec(ws.cell(r, 9).value),
                        valor_total=_dec(ws.cell(r, 10).value),
                        status=_str(ws.cell(r, 11).value) or "EM ABERTO",
                        observacao=_str(ws.cell(r, 12).value),
                    ))
                Romaneio.objects.bulk_create(bulk)
                rc += len(bulk)
                self.stdout.write(f"  ✓ Romaneios {gc}: {len(bulk)}")

            ac = 0
            for gc, kw in [("SF", "abastecimento - sf"), ("SJ", "abastecimento - sj")]:
                ws = _sheet(kw)
                if not ws: continue
                bulk = []
                for r in range(4, ws.max_row + 1):
                    placa = _str(ws.cell(r, 3).value)
                    if not placa: continue
                    bulk.append(Abastecimento(
                        grupo=gc, data_requisicao=_date(ws.cell(r, 1).value),
                        data_abastecimento=_date(ws.cell(r, 5).value),
                        motorista=_mot(_str(ws.cell(r, 2).value), placa),
                        num_requisicao=_str(ws.cell(r, 4).value),
                        qtd_litros=_dec(ws.cell(r, 6).value),
                        vl_unitario=_dec(ws.cell(r, 7).value),
                        valor_posto=_dec(ws.cell(r, 8).value),
                        vl_desconto_total=_dec(ws.cell(r, 10).value),
                        diferenca=_dec(ws.cell(r, 11).value),
                        status=_str(ws.cell(r, 12).value) or "EM ABERTO",
                        observacao=_str(ws.cell(r, 13).value),
                    ))
                Abastecimento.objects.bulk_create(bulk)
                ac += len(bulk)
                self.stdout.write(f"  ✓ Abastecimentos {gc}: {len(bulk)}")

            dc = 0
            for gc, kw in [("SF", "adiantamentos sf"), ("SJ", "adiantamentos sj")]:
                ws = _sheet(kw)
                if not ws: continue
                bulk = []
                for r in range(3, ws.max_row + 1):
                    placa = _str(ws.cell(r, 4).value)
                    if not placa: continue
                    bulk.append(Adiantamento(
                        grupo=gc, data=_date(ws.cell(r, 1).value),
                        descricao=_str(ws.cell(r, 2).value),
                        motorista=_mot(_str(ws.cell(r, 3).value), placa),
                        valor=_dec(ws.cell(r, 5).value),
                        conta_envio=_str(ws.cell(r, 6).value),
                        conta_destino=_str(ws.cell(r, 7).value),
                        status=_str(ws.cell(r, 8).value) or "EM ABERTO",
                        observacao=_str(ws.cell(r, 9).value),
                    ))
                Adiantamento.objects.bulk_create(bulk)
                dc += len(bulk)
                self.stdout.write(f"  ✓ Adiantamentos {gc}: {len(bulk)}")

        wb.close()
        log.romaneios_count = rc
        log.abastecimentos_count = ac
        log.adiantamentos_count = dc
        log.motoristas_count = len(cache)
        log.save()
        self.stdout.write(self.style.SUCCESS(
            f"\n  ✅ {rc} romaneios | {ac} abastecimentos | {dc} adiantamentos | {len(cache)} motoristas\n"
        ))

    def _fail(self, log, msg):
        self.stderr.write(f"  ✗ {msg}")
        log.status, log.erro = "ERRO", msg
        log.save()
